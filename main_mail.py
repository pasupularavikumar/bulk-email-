from pickle import TRUE
import win32com.client as win32
import  time
import openpyxl
from openpyxl import Workbook
import datetime

# Excel file path for data transmission
path_excel = "C:\Users\Admin\Desktop\mail\ABC.xlsx" # file path
wb_obj = openpyxl.load_workbook(path_excel)
sheet_obj = wb_obj.active

#excel file for mails report
wb = Workbook()
sh1 = wb.active
sh1.append(['MAIL_ID','CC','SUBJECT','FILE_ATTACHMENT','REPORT']) 

# Body of the Mail

html_file = 'body.html'
with open(html_file,'r') as file:
    data=file.read()

#send the emails from hear ====>>> range(first_row, last_row+1)

for i in range(2,13):
    Mail_id1 = sheet_obj.cell(row = i, column = 2)  #mail id cell in excel
    CC1 = sheet_obj.cell(row = i, column = 3)       #CC's cell in excel with "," separation
    subject1 = sheet_obj.cell(row = i, column = 4)    #subject cell in excel file  
    #mail_Body = sheet_obj.cell(row=i, column = 5)    # mail body if we need from excel sheet
    Attachment1 = sheet_obj.cell(row = i, column = 6)  #attachments cell with "," separation
    mail_id = str(Mail_id1.value)
    CC = str(CC1.value)
    CC = CC.replace(",", ";")  #cc separation
    subject=str(subject1.value)    
    #mail_body = str(mail_Body.value)
    attachment = str(Attachment1.value)
    attachment1 = attachment.split(",")  #attachments separations
    #sending mail from here
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.createItem(0)
    mail.To = mail_id       # To address
    mail.Subject = subject   # subject
    mail.HTMLBody = data     # Body
    try:
        for x in attachment1:
            mail.Attachments.Add(x)
    except:
        pass
    mail.Cc = CC             # mail cc
    try:
        mail.Send()              # send the mail       
        sh1.append([mail_id , CC, subject, attachment,'Sent']) #mail report if sent
    except:
       sh1.append([mail_id , CC, subject, attachment,'Error']) #mail report for error

    time.sleep(1)            # Time gap for each mail in seconds
# report file name
datetimenow = datetime.datetime.now()
filename = datetimenow.strftime("%Y_%m_%d_%H_%M_%S")+"_mail_report"+".xlsx"
wb.save(filename)

print("all mails are sent") #  final msg 
